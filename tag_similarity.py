from collections import defaultdict
from turtle import color
from types import CodeType, NoneType
import pymysql
import time
import matplotlib.pyplot as plt
import math
import numpy as np
from numpy import dot
from numpy.linalg import norm
from openpyxl import load_workbook
wb=load_workbook("tag_similarity.xlsx")
ws=wb["Sheet1"]
def cos_sim(A, B):
  return dot(A, B)/(norm(A)*norm(B))
db=pymysql.connect(host='localhost',port=3306,user='root',passwd='0000',db='recodb',charset='utf8')
cursor=db.cursor()

tag_index=defaultdict(int)
code_tag_vector=defaultdict(list)
sql=f"select tag,count(tag) from tmp group by tag order by count(tag);"
cursor.execute(sql)
result=cursor.fetchall()
for idx,tag in enumerate(result):
    tag_index[tag[0]]=idx

sql=f"select code,difficulty,id,group_concat(tag separator', ') as tag from tmp group by code order by id;"
cursor.execute(sql)
result=cursor.fetchall()
for idx,res in enumerate(result):
    code=res[0]
    ws.cell(idx+2,1,code)
    ws.cell(1,idx+2,code)
    for i in range(177):
        code_tag_vector[code].append(0)
    tags=res[3].split(',')
    for tag in tags:
        if tag[0]==' ':
            index=tag_index[tag[1:]]
        else:
            index=tag_index[tag]
        code_tag_vector[code][index]=1
    
code1='SEASOR'
arr1=code_tag_vector[code1]

for i in range(2,3351):
    
    
    code2=ws.cell(1,i).value
    
    arr2=code_tag_vector[code2]
    sim=cos_sim(arr1,arr2)
    ws.cell(3350,i,sim)
    ws.cell(i,3350,sim)

wb.save("tag_similarity.xlsx")



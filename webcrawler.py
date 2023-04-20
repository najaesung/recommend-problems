from operator import truediv
from openpyxl import load_workbook
from ast import Pass
from bs4 import BeautifulSoup as bs
from urllib.request import HTTPError
import selenium
import random
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
start=time.time()
options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])
driver=webdriver.Chrome()
wb=load_workbook("problem.xlsx")
ws=wb["algorithms"]
user_sheet=wb["user"]
excel=time.time()
print(f"excel loaded in {excel-start:.5f} sec")
row =2869
page = 0  #max:106
dictionary={}
problem_table={}
for i in range(2,3492):
        name=user_sheet.cell(1,i).value
        problem_table[name]=i
list_html='https://www.codechef.com/practice?page=%d&limit=20&sort_by=difficulty_rating&sort_order=asc&search=&start_rating=0&end_rating=5000&topic=&tags=&group=all'

def login_codechef():
    ID='jsr123'
    Password='Dhwltla18~'

    

    headers = {'User-Agent':'Chrome/66.0.3359.181'}
    url='https://www.codechef.com/login?destination=/'
    driver.get(url)
    driver.implicitly_wait(5)
    driver.find_element_by_xpath("/html/body/section/div[2]/div/div[3]/div[2]/form/div/div[1]/div/div/input").send_keys(ID)
    driver.find_element_by_xpath("/html/body/section/div[2]/div/div[3]/div[2]/form/div/div[2]/div/div[2]/input").send_keys(Password)
    driver.find_element_by_xpath("/html/body/section/div[2]/div/div[3]/div[2]/form/div/div[3]/input").click()
    driver.implicitly_wait(5)

def scrape():
    global row
    for j in range(143,179):
        driver.get(list_html%j)
        time.sleep(5)
        if(j==0):
            for i in range(0,28):
                html=driver.page_source
                soup=bs(html,'html.parser')
                name=soup.select_one('#MUIDataTableBodyRow-%d > td:nth-child(2) > div:nth-child(2) > a'%i).get_text()
                addr=soup.select_one('#MUIDataTableBodyRow-%d > td:nth-child(1) > div:nth-child(2)'%i).get_text()
                driver.get("https://www.codechef.com/submit-v2/"+addr)
                driver.implicitly_wait(5)
                time.sleep(2)
                html=driver.page_source
                soup=bs(html,'html.parser')
                tags=soup.select('#root > div > div.ContainerBox_pageContainer__1zfPw > div > div > div.TopBanner_problem-banner__container__1vpej > div.TopBanner_content-ads__container__18nn1 > div.TopBanner_content__wrapper__1RoeN > div.TopBanner_ratings-tags__container__3xrrb > div.TopBanner_tags-list__container__2Zv5d > div.TopBanner_tag-list-map__box__1-RqG > div')
                p=soup.select('#problem-statement')[0].get_text()
                a=p.find("Input")
                problem=p[7:a]
                ws.cell(row,1,addr)
                ws.cell(row,2,name)
                ws.cell(row,3,problem)
                for idx,k in enumerate(tags):
                    ws.cell(row,4+idx,k.get_text())
                row+=1
                wb.save("problem.xlsx")
                print("name: %s"%name)
                print("i: %d"%i)
                print("j: %d"%j)
                print("row: %d\n"%row)
                driver.back()
                time.sleep(2)
        else:
            for i in range(0,20):
                html=driver.page_source
                soup=bs(html,'html.parser')
                name=soup.select_one('#MUIDataTableBodyRow-%d > td:nth-child(2) > div:nth-child(2) > a'%i).get_text()
                addr=soup.select_one('#MUIDataTableBodyRow-%d > td:nth-child(1) > div:nth-child(2)'%i).get_text()
                driver.get("https://www.codechef.com/submit-v2/"+addr)
                driver.implicitly_wait(5)
                time.sleep(2)
                html=driver.page_source
                soup=bs(html,'html.parser')
                tags=soup.select('#root > div > div.ContainerBox_pageContainer__1zfPw > div > div > div.TopBanner_problem-banner__container__1vpej > div.TopBanner_content-ads__container__18nn1 > div.TopBanner_content__wrapper__1RoeN > div.TopBanner_ratings-tags__container__3xrrb > div.TopBanner_tags-list__container__2Zv5d > div.TopBanner_tag-list-map__box__1-RqG > div')
                p=soup.select('#problem-statement')[0].get_text()
                a=p.find("Input")
                problem=p[7:a]
                ws.cell(row,1,addr)
                ws.cell(row,2,name)
                ws.cell(row,3,problem)
                for idx,k in enumerate(tags):
                    ws.cell(row,4+idx,k.get_text())
                row+=1
                wb.save("problem.xlsx")
                print("name: %s"%name)
                print("i: %d"%i)
                print("j: %d"%j)
                print("row: %d\n"%row)
                driver.back()
                time.sleep(2)


def user_scrape():
    row=27108
    
    for i in range(9,3491):
        code=ws.cell(i,1).value
        url='https://www.codechef.com/status/'+code
        driver.get(url)
        driver.implicitly_wait(5)
        html=driver.page_source
        soup=bs(html,'html.parser')
        m=soup.select_one('#content > div > div > table > tbody > tr > td:nth-child(2) > div').get_text()
        k=m.find('of')
        mx=m[k+3:]
        max=int(mx)
        if(max>1000):
            max=1000
        for k in range(1,max-1):
            
            driver.get(url+"?page=%d"%k)
            driver.implicitly_wait(5)
                
            html=driver.page_source
            soup=bs(html,'html.parser')
            
            for j in range(1,26):
                
                try:
                    u=soup.select_one('#content > div > div > div.tablebox-section.l-float > table > tbody > tr:nth-child(%d) > td:nth-child(3) > a'%j).get_text()
                except:
                    driver.refresh()
                    driver.implicitly_wait(5)
                    time.sleep(2)
                    html=driver.page_source
                    soup=bs(html,'html.parser')
                    u=soup.select_one('#content > div > div > div.tablebox-section.l-float > table > tbody > tr:nth-child(%d) > td:nth-child(3) > a'%j).get_text()
                
                img=soup.select_one('#content > div > div > div.tablebox-section.l-float > table > tbody > tr:nth-child(%d) > td:nth-child(4) > span > img'%j).get('src')
                if(img=='https://cdn.codechef.com/misc/tick-icon.gif'):
                    solve=True
                else:
                    solve=False

                a=u.find("★")
                user_name=u[a+1:]
                find=user_name in dictionary

                if(find==False):
                    dictionary[user_name]=row
                    user_sheet.cell(row,1,user_name)
                    if(solve==True):
                        user_sheet.cell(row,i+1,"O")
                    else:
                        user_sheet.cell(row,i+1,"X")
                    row+=1
                   
                else:
                    r=dictionary[user_name]
                    if(solve==True):
                        user_sheet.cell(r,i+1,"O")
                    else:
                        if(user_sheet.cell(r,i+1).value!="O"):
                            user_sheet.cell(r,i+1,"X")
                
            if(k%50==0 or k==max-2):
                print("i:%d\tk:%d\trow:%d"%(i,k,row))
                wb.save("problem.xlsx")
                print('excel saved')

def user_base_scrape():
    
    f=open('surp.csv','a',encoding='UTF-8')
    newf=open('new_problems.txt','a',encoding='UTF-8')
    for i in range(165208,207788):
        prob_score={}
        user_name=user_sheet.cell(i,1).value
        if(user_name==None):
            print('vjudge account\t i:%d'%i)
        else:
            url='https://www.codechef.com/users/'+user_name
            driver.get(url)
            driver.implicitly_wait(5)
            time.sleep(1)
            driver.execute_script("window.scrollTo(0, 1200)")
            driver.implicitly_wait(5)
            time.sleep(0.3)
            html=driver.page_source
            soup=bs(html,'html.parser')
            try:
                error=soup.select_one('#maintable > tbody > tr > td > div > div > div > div > p').get_text()
                print('user deleted')
            except:
                try:
                    m=soup.select_one('#loader > div').get_text()
                except:
                    driver.refresh()
                    driver.implicitly_wait(5)
                    time.sleep(2)
                    html=driver.page_source
                    soup=bs(html,'html.parser')
                    try:
                        m=soup.select_one('#loader > div').get_text()
                    except:
                        print('HIHIHIHI IN EXCEPT!')
                        continue
                k=m.find('of')
                if(k==-1):
                    continue
                mx=m[k+3:]
                max=int(mx)
                for j in range(1,max+1):
                    html=driver.page_source
                    soup=bs(html,'html.parser')
                    for k in range(1,13):
                        try:
                            problem_code=soup.select_one('#rankContentDiv > div:nth-child(1) > table > tbody > tr:nth-child(%d) > td:nth-child(2) > a'%k).get_text()
                        except:
                            continue
                        
                        if(problem_code not in problem_table):
                            newf.write(problem_code+'\n')
                        img=soup.select_one('#rankContentDiv > div:nth-child(1) > table > tbody > tr:nth-child(%d) > td:nth-child(3) > span > img'%k).get('src')
                            
                        if(img=='https://cdn.codechef.com/misc/tick-icon.gif'):
                            solve=True
                        else:
                            solve=False
                         
                        if(solve==True):
                            #user_sheet.cell(i,problem_table[problem_code],'O')
                            prob_score[problem_code]='100'
                        else:
                            if(prob_score.get(problem_code)!='100'):
                                prob_score[problem_code]='0'
                            #if(user_sheet.cell(i,problem_table[problem_code]).value!="O"):
                            #   user_sheet.cell(i,problem_table[problem_code],'X')
                        
                                    
                    
                    try:
                        driver.find_element(By.XPATH,'//*[@id="rankContentDiv"]/table/tbody/tr/td[3]/a').click() 
                        driver.implicitly_wait(5)  
                        
                        
                    except:
                        if(j==max):
                            print('last page done')
                            continue
                        print('not clicked')
                        driver.execute_script("window.scrollTo(0, 1200)")
                        driver.implicitly_wait(5)
                        time.sleep(1)
                        driver.find_element(By.XPATH,'//*[@id="rankContentDiv"]/table/tbody/tr/td[3]/a').click() 
                        driver.implicitly_wait(5)  
                        
                
                for code in prob_score:
                    f.write(user_name + ',' + code + ',' + prob_score[code]+'\n')
                print('user_id: %s\ti: %d\t max: %d'%(user_name,i,max))
                now=time.time()
                print(f"running time: {(now-start)//3600} hours {((now-start)//60)%60} minutes {(now-start)%60:.5f} secs count:{i-165207}")
        
                
    f.close()
    newf.close()

                    
                
                
def user_exists(user_name):
    
    for idx,row in enumerate(user_sheet.iter_rows(min_row=1)):
        if row[0].value == user_name:
            return idx+1
    return False
        
        
def main():
    user_base_scrape()
    #login_codechef()
    #scrape()
    #user_scrape()

if __name__=="__main__":
    main()



